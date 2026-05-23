import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime

FILE = r"D:\PowerChina\Anexo 01-LPA.xlsm"

wb = load_workbook(FILE, read_only=True, keep_vba=True, data_only=True)
ws = wb["LPA CONSTR."]

print("=== ABA: LPA CONSTR. ===")
print("Lendo todas as linhas...\n")

all_rows = []
for row in ws.iter_rows(values_only=True):
    all_rows.append(row)

print(f"Total de linhas: {len(all_rows)}")
print(f"Total de colunas: {max(len(r) for r in all_rows)}")

# Print first 30 rows to understand structure
print("\n--- PRIMEIRAS 30 LINHAS (completas) ---")
for i, row in enumerate(all_rows[:30]):
    non_none = {j: v for j, v in enumerate(row) if v is not None}
    print(f"Linha {i+1}: {non_none}")

wb.close()
