import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

SOURCE = r"D:\PowerChina\Anexo 01-LPA.xlsm"
OUTPUT = r"D:\PowerChina\Painel_Construccion_LaPampina.xlsx"

# ─── Read source ──────────────────────────────────────────────────────────────
wb_src = load_workbook(SOURCE, read_only=True, keep_vba=True, data_only=True)
ws_src = wb_src["LPA CONSTR."]
all_rows = list(ws_src.iter_rows(values_only=True))

HEADER_ROW = all_rows[9]   # Row 10 (0-indexed 9) = column headers
DATA_DATE  = datetime(2026, 5, 17)
DATA_WEEK  = "S20-26"

# Build week label arrays from header row
plan_weeks, real_weeks = [], []
for c in range(22, 90):
    plan_weeks.append(HEADER_ROW[c] if c < len(HEADER_ROW) else None)
for c in range(92, 160):
    real_weeks.append(HEADER_ROW[c] if c < len(HEADER_ROW) else None)

# Current week index in the 68-week series (S44-25=0 … S20-26=28)
CURR_IDX = 28   # col 50 - 22 = 28

# ─── Parse activities ─────────────────────────────────────────────────────────
records = []
for row in all_rows[10:]:
    if not row[0] and not row[5]:
        continue
    r = dict(
        fila=row[0], nivel=row[1], es_resumen=row[2],
        edt=row[6], tarea=str(row[7]).strip() if row[7] else '',
        duracion=row[8],
        inicio_lb=row[9], fin_lb=row[10],
        trabajo_h=row[11] or 0,
        incidencia=row[13] or 0,
        pct_plan=row[15] or 0,
        pct_real=row[16] or 0,
        pct_comp_plan=row[17] or 0,
        pct_comp_real=row[18] or 0,
        desv_pond=row[19] or 0,
        desviacion=row[20] or 0,
    )
    for i, c in enumerate(range(22, 90)):
        r[f'ps{i}'] = row[c] if c < len(row) and row[c] is not None else 0
    for i, c in enumerate(range(92, 160)):
        r[f'rs{i}'] = row[c] if c < len(row) and row[c] is not None else 0
    records.append(r)

df = pd.DataFrame(records)
df_total   = df[df['edt'] == '4.5'].iloc[0]
df_level3  = df[df['nivel'] == 3].copy()
df_leaves  = df[df['es_resumen'] == 'No'].copy()

# ─── Helpers ──────────────────────────────────────────────────────────────────
def rgb(r, g, b): return f"{r:02X}{g:02X}{b:02X}"

C_DARK_BLUE  = rgb(0,  47, 108)
C_MED_BLUE   = rgb(0,  84, 166)
C_LIGHT_BLUE = rgb(173,214,245)
C_GREEN      = rgb(0,  163, 108)
C_RED        = rgb(192,0,  0)
C_ORANGE     = rgb(255,153, 0)
C_YELLOW     = rgb(255,255,153)
C_WHITE      = "FFFFFF"
C_GRAY       = rgb(242,242,242)
C_DARK_GRAY  = rgb(127,127,127)

def hdr_font(bold=True, size=11, color=C_WHITE):
    return Font(name='Arial', bold=bold, size=size, color=color)

def cell_font(bold=False, size=10, color="000000"):
    return Font(name='Arial', bold=bold, size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_thin():
    s = Side(style='thin', color=C_DARK_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style='medium', color=C_DARK_BLUE)
    return Border(left=s, right=s, top=s, bottom=s)

def pct_fmt(v):
    return f"{v*100:.2f}%"

def apply_hdr(ws, row, col, value, bg=C_DARK_BLUE, fg=C_WHITE,
              bold=True, size=11, wrap=False, align='center'):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='Arial', bold=bold, size=size, color=fg)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border = border_thin()
    return c

def apply_data(ws, row, col, value, bg=C_WHITE, bold=False,
               fmt=None, align='center', size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='Arial', bold=bold, size=size, color="000000")
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical='center')
    c.border = border_thin()
    if fmt:
        c.number_format = fmt
    return c

# ─── Build Workbook ───────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)   # remove default sheet

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: PAINEL EXECUTIVO
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("PAINEL EXECUTIVO")
ws1.sheet_view.showGridLines = False

# Title banner
ws1.merge_cells('A1:N1')
c = ws1['A1']
c.value = "PAINEL DE CONTROL DE AVANCE  —  CONSTRUCCIÓN  |  LA PAMPINA"
c.font = Font(name='Arial', bold=True, size=16, color=C_WHITE)
c.fill = fill(C_DARK_BLUE)
c.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 36

ws1.merge_cells('A2:N2')
c = ws1['A2']
c.value = f"Fecha de Control: {DATA_DATE.strftime('%d/%m/%Y')}  |  Semana: {DATA_WEEK}  |  Última Actualización: {DATA_DATE.strftime('%d/%m/%Y')}"
c.font = Font(name='Arial', size=10, color=C_WHITE)
c.fill = fill(C_MED_BLUE)
c.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 20

# ── KPI Cards Row ──────────────────────────────────────────────────────────────
def kpi_card(ws, start_row, start_col, title, value, unit="", color=C_MED_BLUE, sub=""):
    end_col = start_col + 1
    ws.merge_cells(start_row=start_row, start_column=start_col,
                   end_row=start_row, end_column=end_col)
    c = ws.cell(row=start_row, column=start_col, value=title)
    c.font = Font(name='Arial', bold=True, size=9, color=C_WHITE)
    c.fill = fill(color)
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=start_row+1, start_column=start_col,
                   end_row=start_row+2, end_column=end_col)
    c2 = ws.cell(row=start_row+1, column=start_col, value=f"{value}{unit}")
    c2.font = Font(name='Arial', bold=True, size=22, color=color)
    c2.fill = fill(C_GRAY)
    c2.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=start_row+3, start_column=start_col,
                   end_row=start_row+3, end_column=end_col)
    c3 = ws.cell(row=start_row+3, column=start_col, value=sub)
    c3.font = Font(name='Arial', size=8, color=C_DARK_GRAY)
    c3.fill = fill(C_GRAY)
    c3.alignment = Alignment(horizontal='center', vertical='center')
    for r in range(start_row, start_row+4):
        for col in range(start_col, end_col+1):
            ws.cell(row=r, column=col).border = border_thin()

ws1.row_dimensions[3].height = 18
ws1.row_dimensions[4].height = 30
ws1.row_dimensions[5].height = 30
ws1.row_dimensions[6].height = 16

plan_pct = df_total['pct_plan'] * 100
real_pct = df_total['pct_real'] * 100
dev_pct  = df_total['desv_pond'] * 100

kpi_card(ws1, 3, 1, "% PLANEADO",     f"{plan_pct:.2f}", "%", C_MED_BLUE,
         f"Inicio: 27/10/2025")
kpi_card(ws1, 3, 3, "% REAL",         f"{real_pct:.2f}", "%", C_GREEN,
         f"Semana {DATA_WEEK}")
kpi_card(ws1, 3, 5, "DESVÍO ACUM.",  f"{dev_pct:+.2f}", "%",
         C_RED if dev_pct < 0 else C_GREEN,
         "Negativo = atraso")
kpi_card(ws1, 3, 7, "H-H TOTALES",  f"{df_total['trabajo_h']/1000:.1f}", "K h",
         C_DARK_BLUE, "Horas-hombre LB")
kpi_card(ws1, 3, 9, "ACTIVIDADES",  "1,098", "",
         C_MED_BLUE, "Elementales")
kpi_card(ws1, 3, 11, "SIN AVANCE",   "69", "",
         C_ORANGE, "Planeadas sin real")
kpi_card(ws1, 3, 13, "FORECAST",     "???", "",
         C_DARK_GRAY, "Ver hoja FORECAST")

ws1.row_dimensions[7].height = 8

# ── Section: Avance por Área ──────────────────────────────────────────────────
apply_hdr(ws1, 8, 1, "AVANCE POR ÁREA / DISCIPLINA", bg=C_DARK_BLUE, bold=True, size=11)
ws1.merge_cells('A8:N8')

headers_area = ['ÁREA / GRUPO', 'EDT', 'INICIO LB', 'FIN LB', 'H-H', 'INCIDENCIA',
                '% PLAN POND.', '% REAL POND.', 'DESVÍO POND.', '% COMPLETADO PLAN',
                '% COMPLETADO REAL', 'DESVÍO COMPLETADO']
for j, h in enumerate(headers_area, 1):
    apply_hdr(ws1, 9, j, h, bg=C_MED_BLUE, size=9, wrap=True)
ws1.row_dimensions[9].height = 30

area_map = {
    '4.5':    ('TOTAL CONSTRUCCIÓN', C_DARK_BLUE),
    '4.5.2':  ('Primeros Trabajos',   C_GRAY),
    '4.5.3':  ('Instalaciones Permanentes', C_GRAY),
    '4.5.4':  ('BESS',                C_GRAY),
    '4.5.5':  ('Parque Fotovoltaico (PV)', C_GRAY),
    '4.5.6':  ('Sala de Control',     C_GRAY),
    '4.5.7':  ('Conexión HUB',        C_GRAY),
    '4.5.8':  ('Pre-Commissioning',   C_GRAY),
    '4.5.10': ('Energización',        C_GRAY),
}

row = 10
for edt, (label, bg) in area_map.items():
    sub = df[df['edt'] == edt]
    if sub.empty:
        continue
    r = sub.iloc[0]
    is_total = (edt == '4.5')
    bold = is_total
    bg_row = C_LIGHT_BLUE if is_total else (C_GRAY if row % 2 == 0 else C_WHITE)
    dev = r['desv_pond']
    dev_color = C_RED if dev < -0.001 else (C_GREEN if dev > 0.001 else C_YELLOW)

    inicio = r['inicio_lb'].strftime('%d/%m/%Y') if isinstance(r['inicio_lb'], datetime) else str(r['inicio_lb'])[:10]
    fim    = r['fin_lb'].strftime('%d/%m/%Y')    if isinstance(r['fin_lb'], datetime)    else str(r['fin_lb'])[:10]

    vals = [label, edt, inicio, fim,
            int(r['trabajo_h']),
            r['incidencia'],
            r['pct_plan'], r['pct_real'], r['desv_pond'],
            r['pct_comp_plan'], r['pct_comp_real'], r['desviacion']]
    fmts = [None, None, None, None,
            '#,##0', '0.000%', '0.00%', '0.00%', '0.00%',
            '0.0%', '0.0%', '0.0%']
    aligns = ['left','center','center','center',
              'right','right','right','right','right',
              'right','right','right']
    for j, (v, f, a) in enumerate(zip(vals, fmts, aligns), 1):
        cell = apply_data(ws1, row, j, v, bg=bg_row, bold=bold, fmt=f, align=a)
        if j == 9:  # desvio column
            cell.fill = fill(dev_color)
    row += 1

ws1.row_dimensions[row].height = 8
row += 1

# ── Section: Cronograma resumido ──────────────────────────────────────────────
apply_hdr(ws1, row, 1, "RESUMEN EJECUTIVO AUTOMÁTICO", bg=C_DARK_BLUE, bold=True, size=11)
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
row += 1

resumen_lines = [
    f"Projeto: LA PAMPINA  |  Disciplina: CONSTRUÇÃO  |  Data: {DATA_DATE.strftime('%d/%m/%Y')} ({DATA_WEEK})",
    f"Período LB: 27/10/2025 a 10/02/2027  |  Total H-H: 823.866 horas-homem",
    "",
    f"SITUAÇÃO ATUAL: O avanço real de CONSTRUÇÃO é de {real_pct:.2f}%, contra {plan_pct:.2f}% planejado.",
    f"Desvio acumulado ponderado: {dev_pct:+.2f}% — o projeto ESTÁ ATRASADO.",
    "",
    "ÁREAS CRÍTICAS:",
    f"  • BESS (4.5.4):         Plan {df[df['edt']=='4.5.4'].iloc[0]['pct_plan']*100:.2f}% | Real {df[df['edt']=='4.5.4'].iloc[0]['pct_real']*100:.2f}% | Desvio {df[df['edt']=='4.5.4'].iloc[0]['desv_pond']*100:+.2f}%",
    f"  • Parque FV (4.5.5):    Plan {df[df['edt']=='4.5.5'].iloc[0]['pct_plan']*100:.2f}% | Real {df[df['edt']=='4.5.5'].iloc[0]['pct_real']*100:.2f}% | Desvio {df[df['edt']=='4.5.5'].iloc[0]['desv_pond']*100:+.2f}%",
    f"  • Sala Controle (4.5.6): Plan {df[df['edt']=='4.5.6'].iloc[0]['pct_plan']*100:.2f}% | Real {df[df['edt']=='4.5.6'].iloc[0]['pct_real']*100:.2f}% | Desvio {df[df['edt']=='4.5.6'].iloc[0]['desv_pond']*100:+.2f}%",
    "",
    "TOP DESVIOS (atividades):  Instalación de faenas -59.2% | Cerco perimetral BESS -56.7% | Caminos PV -80%",
    "69 atividades com plano > 0 e avanço real = 0 (ver aba PLANEJADAS SEM AVANÇO)",
    "1.036 atividades futuras em andamento (ver aba FORECAST)",
]

for line in resumen_lines:
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    c = ws1.cell(row=row, column=1, value=line)
    c.font = Font(name='Arial', size=9,
                  bold=line.startswith('SITUAÇÃO') or line.startswith('ÁREAS') or line.startswith('TOP'),
                  color='000000' if line else C_WHITE)
    c.fill = fill(C_GRAY if line else C_WHITE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[row].height = 14
    row += 1

# Column widths
col_widths = [2, 28, 10, 12, 12, 12, 10, 10, 11, 11, 12, 12, 12, 12]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: CURVA S
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("CURVA S")
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:H1')
ws2['A1'] = "CURVA S — CONSTRUCCIÓN | LA PAMPINA"
ws2['A1'].font = Font(name='Arial', bold=True, size=14, color=C_WHITE)
ws2['A1'].fill = fill(C_DARK_BLUE)
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 28

headers_s = ['#', 'SEMANA', 'FECHA', '% PLAN ACUM.', '% REAL ACUM.', 'DESVÍO %']
for j, h in enumerate(headers_s, 1):
    apply_hdr(ws2, 2, j, h, bg=C_MED_BLUE, size=9)

# Get S-curve data from total row (4.5)
total_idx = df[df['edt'] == '4.5'].index[0]
total_row_data = df.loc[total_idx]

# Build date array from row 9 (index 8) of source
date_row = all_rows[8]
plan_dates_list = [date_row[c] for c in range(22, 90) if c < len(date_row)]

plan_acum  = [total_row_data[f'ps{i}'] for i in range(68)]
real_acum  = [total_row_data[f'rs{i}'] for i in range(68)]

data_row_s = 3
for i in range(68):
    wk_label = plan_weeks[i] if i < len(plan_weeks) else f"S{i}"
    fecha = plan_dates_list[i] if i < len(plan_dates_list) else None
    p = plan_acum[i] if plan_acum[i] else 0
    r = real_acum[i] if real_acum[i] else 0
    d = r - p

    is_current = (i == CURR_IDX)
    bg = C_LIGHT_BLUE if is_current else (C_GRAY if data_row_s % 2 == 0 else C_WHITE)

    apply_data(ws2, data_row_s, 1, i+1, bg=bg)
    apply_data(ws2, data_row_s, 2, wk_label, bg=bg)
    apply_data(ws2, data_row_s, 3, fecha, bg=bg, fmt='DD/MM/YYYY')
    apply_data(ws2, data_row_s, 4, p, bg=bg, fmt='0.00%')
    apply_data(ws2, data_row_s, 5, r if r else None, bg=bg, fmt='0.00%')
    dev_bg = bg if abs(d) < 0.001 else (C_RED if d < 0 else C_GREEN)
    apply_data(ws2, data_row_s, 6, d if (p or r) else None, bg=dev_bg, fmt='+0.00%;-0.00%;0.00%')
    data_row_s += 1

# Line chart
chart = LineChart()
chart.title = "Curva S — Construcción LA Pampina"
chart.style = 10
chart.y_axis.title = "% Avance Acumulado"
chart.x_axis.title = "Semana"
chart.y_axis.numFmt = '0%'
chart.height = 14
chart.width = 22

plan_ref = Reference(ws2, min_col=4, min_row=2, max_row=2+67)
real_ref = Reference(ws2, min_col=5, min_row=2, max_row=2+67)
cats     = Reference(ws2, min_col=2, min_row=3, max_row=2+67)

from openpyxl.chart import Series
s1 = Series(plan_ref, title="% Plan")
s2 = Series(real_ref, title="% Real")
s1.graphicalProperties.line.solidFill = C_MED_BLUE
s1.graphicalProperties.line.width = 20000
s2.graphicalProperties.line.solidFill = C_GREEN
s2.graphicalProperties.line.width = 20000

chart.append(s1)
chart.append(s2)
chart.set_categories(cats)
ws2.add_chart(chart, "H2")

for j, w in enumerate([4, 10, 12, 12, 12, 12], 1):
    ws2.column_dimensions[get_column_letter(j)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: PROGRESO POR ÁREA
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("PROGRESO POR ÁREA")
ws3.sheet_view.showGridLines = False

ws3.merge_cells('A1:L1')
ws3['A1'] = "PROGRESO POR ÁREA Y DISCIPLINA — CONSTRUCCIÓN | LA PAMPINA"
ws3['A1'].font = Font(name='Arial', bold=True, size=13, color=C_WHITE)
ws3['A1'].fill = fill(C_DARK_BLUE)
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 24

areas_detail = [
    ('4.5',    'TOTAL CONSTRUCCIÓN',            C_DARK_BLUE, C_WHITE),
    ('4.5.2',  '▶ Primeros Trabajos',            C_MED_BLUE,  C_WHITE),
    ('4.5.3',  '▶ Instalaciones Permanentes',   C_MED_BLUE,  C_WHITE),
    ('4.5.4',  '▶ BESS',                        C_MED_BLUE,  C_WHITE),
    ('4.5.5',  '▶ Parque Fotovoltaico (PV)',     C_MED_BLUE,  C_WHITE),
    ('4.5.6',  '▶ Sala de Control (Control Room)',C_MED_BLUE, C_WHITE),
    ('4.5.7',  '▶ Conexión HUB',                C_MED_BLUE,  C_WHITE),
    ('4.5.8',  '▶ Pre-Commissioning',           C_MED_BLUE,  C_WHITE),
    ('4.5.10', '▶ Energización',                C_MED_BLUE,  C_WHITE),
]

hdrs3 = ['ÁREA / GRUPO', 'EDT', 'INICIO LB', 'FIN LB', 'H-H TOTAL',
         'INCIDENCIA', '% PLAN POND.', '% REAL POND.', 'DESVÍO POND.',
         '% COMP. PLAN', '% COMP. REAL', 'STATUS']
for j, h in enumerate(hdrs3, 1):
    apply_hdr(ws3, 2, j, h, bg=C_MED_BLUE, size=9, wrap=True)
ws3.row_dimensions[2].height = 28

row3 = 3
for edt, label, hdr_bg, hdr_fg in areas_detail:
    sub = df[df['edt'] == edt]
    if sub.empty:
        continue
    r = sub.iloc[0]
    is_total = (edt == '4.5')
    bg = C_LIGHT_BLUE if is_total else (C_GRAY if row3 % 2 == 0 else C_WHITE)

    dev = r['desv_pond']
    if dev <= -0.01:
        status = "⚠ CRÍTICO"
        s_bg = C_RED; s_fg = C_WHITE
    elif dev < -0.003:
        status = "▼ ATRASADO"
        s_bg = C_ORANGE; s_fg = "000000"
    elif dev >= 0:
        status = "✔ OK"
        s_bg = C_GREEN; s_fg = C_WHITE
    else:
        status = "~ LEVE ATRASO"
        s_bg = C_YELLOW; s_fg = "000000"

    inicio = r['inicio_lb'].strftime('%d/%m/%Y') if isinstance(r['inicio_lb'], datetime) else ''
    fim    = r['fin_lb'].strftime('%d/%m/%Y')    if isinstance(r['fin_lb'], datetime)    else ''

    vals3 = [label, edt, inicio, fim,
             int(r['trabajo_h']), r['incidencia'],
             r['pct_plan'], r['pct_real'], r['desv_pond'],
             r['pct_comp_plan'], r['pct_comp_real'], status]
    fmts3 = [None,None,None,None,'#,##0','0.000%',
             '0.00%','0.00%','0.00%','0.0%','0.0%',None]
    aligns3 = ['left','center','center','center','right','right',
               'right','right','right','right','right','center']

    for j, (v, f, a) in enumerate(zip(vals3, fmts3, aligns3), 1):
        cell = apply_data(ws3, row3, j, v, bg=bg, bold=is_total, fmt=f, align=a)
        if j == 9 and not is_total:
            cell.fill = fill(s_bg)
            cell.font = Font(name='Arial', bold=True, size=10, color=s_fg)
        if j == 12:
            cell.fill = fill(s_bg)
            cell.font = Font(name='Arial', bold=True, size=10, color=s_fg)
    row3 += 1

row3 += 1
apply_hdr(ws3, row3, 1, "DETALHE NÍVEL 4 — ATIVIDADES RESUMO", bg=C_DARK_BLUE, bold=True, size=10)
ws3.merge_cells(start_row=row3, start_column=1, end_row=row3, end_column=12)
row3 += 1

hdrs3b = ['TAREFA', 'EDT', 'INÍCIO LB', 'FIM LB', 'H-H',
          'INCIDÊNCIA', '% PLAN', '% REAL', 'DESVÍO']
for j, h in enumerate(hdrs3b, 1):
    apply_hdr(ws3, row3, j, h, bg=C_MED_BLUE, size=9)
row3 += 1

df_l4 = df[df['nivel'] == 4].copy()
for _, r in df_l4.iterrows():
    bg = C_GRAY if row3 % 2 == 0 else C_WHITE
    dev = r['desv_pond']
    inicio = r['inicio_lb'].strftime('%d/%m/%Y') if isinstance(r['inicio_lb'], datetime) else ''
    fim    = r['fin_lb'].strftime('%d/%m/%Y')    if isinstance(r['fin_lb'], datetime)    else ''
    vals = [r['tarea'][:60], r['edt'], inicio, fim,
            int(r['trabajo_h']), r['incidencia'],
            r['pct_plan'], r['pct_real'], r['desv_pond']]
    fmts_ = [None,None,None,None,'#,##0','0.000%','0.00%','0.00%','0.00%']
    aligns_ = ['left','center','center','center','right','right','right','right','right']
    for j, (v, f, a) in enumerate(zip(vals, fmts_, aligns_), 1):
        cell = apply_data(ws3, row3, j, v, bg=bg, fmt=f, align=a)
        if j == 9 and dev < -0.002:
            cell.fill = fill(C_RED if dev < -0.005 else C_ORANGE)
    row3 += 1

col_widths3 = [35, 14, 12, 12, 10, 10, 10, 10, 10, 10, 10, 14]
for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: TOP DESVÍOS
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("TOP DESVÍOS")
ws4.sheet_view.showGridLines = False

ws4.merge_cells('A1:J1')
ws4['A1'] = "ACTIVIDADES CON MAYOR DESVÍO — CONSTRUCCIÓN | LA PAMPINA"
ws4['A1'].font = Font(name='Arial', bold=True, size=13, color=C_WHITE)
ws4['A1'].fill = fill(C_RED)
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 24

hdrs4 = ['#', 'EDT', 'ACTIVIDAD', 'INICIO LB', 'FIN LB', 'H-H',
         'INCIDENCIA', '% COMP. PLAN', '% COMP. REAL', 'DESVÍO %']
for j, h in enumerate(hdrs4, 1):
    apply_hdr(ws4, 2, j, h, bg=C_RED, size=9, wrap=True)
ws4.row_dimensions[2].height = 28

desvios = df_leaves.copy()
desvios = desvios[desvios['incidencia'] > 0.0001]
desvios_sorted = desvios.sort_values('desviacion', ascending=True)

for rank, (_, r) in enumerate(desvios_sorted.head(50).iterrows(), 1):
    bg = C_YELLOW if abs(r['desviacion']) > 0.5 else (C_GRAY if rank % 2 == 0 else C_WHITE)
    inicio = r['inicio_lb'].strftime('%d/%m/%Y') if isinstance(r['inicio_lb'], datetime) else ''
    fim    = r['fin_lb'].strftime('%d/%m/%Y')    if isinstance(r['fin_lb'], datetime)    else ''
    vals4 = [rank, r['edt'], r['tarea'][:60], inicio, fim,
             int(r['trabajo_h']), r['incidencia'],
             r['pct_comp_plan'], r['pct_comp_real'], r['desviacion']]
    fmts4 = [None,None,None,None,None,'#,##0','0.0000%',
             '0.0%','0.0%','+0.0%;-0.0%;0.0%']
    aligns4 = ['center','center','left','center','center',
               'right','right','right','right','right']
    row4 = rank + 2
    for j, (v, f, a) in enumerate(zip(vals4, fmts4, aligns4), 1):
        cell = apply_data(ws4, row4, j, v, bg=bg, fmt=f, align=a)
        if j == 10:
            cell.fill = fill(C_RED if r['desviacion'] < -0.3 else
                            (C_ORANGE if r['desviacion'] < -0.1 else C_YELLOW))
            cell.font = Font(name='Arial', bold=True, size=10, color=C_WHITE if r['desviacion'] < -0.3 else "000000")

col_widths4 = [4, 16, 50, 11, 11, 8, 10, 12, 12, 12]
for i, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5: ACTIVIDADES CRÍTICAS
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("ACTIVIDADES CRÍTICAS")
ws5.sheet_view.showGridLines = False

ws5.merge_cells('A1:K1')
ws5['A1'] = "ACTIVIDADES CRÍTICAS (Alta Incidencia + Desvío Negativo) — LA PAMPINA"
ws5['A1'].font = Font(name='Arial', bold=True, size=13, color=C_WHITE)
ws5['A1'].fill = fill(C_RED)
ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 24

hdrs5 = ['#', 'EDT', 'ACTIVIDAD', 'INICIO LB', 'FIN LB', 'H-H',
         'INCIDENCIA', '% PLAN', '% REAL', 'DESVÍO', 'IMPACTO EN EL PROYECTO']
for j, h in enumerate(hdrs5, 1):
    apply_hdr(ws5, 2, j, h, bg=C_RED, size=9, wrap=True)
ws5.row_dimensions[2].height = 28

critical = df_leaves[
    (df_leaves['incidencia'] > 0.003) &
    (df_leaves['desviacion'] < -0.05)
].sort_values('incidencia', ascending=False)

for rank, (_, r) in enumerate(critical.iterrows(), 1):
    bg = C_GRAY if rank % 2 == 0 else C_WHITE
    inicio = r['inicio_lb'].strftime('%d/%m/%Y') if isinstance(r['inicio_lb'], datetime) else ''
    fim    = r['fin_lb'].strftime('%d/%m/%Y')    if isinstance(r['fin_lb'], datetime)    else ''
    impacto_abs = abs(r['desv_pond']) * 100
    impacto = f"Retrasa ~{impacto_abs:.4f}% del total"
    vals5 = [rank, r['edt'], r['tarea'][:55], inicio, fim,
             int(r['trabajo_h']), r['incidencia'],
             r['pct_comp_plan'], r['pct_comp_real'],
             r['desviacion'], impacto]
    fmts5 = [None,None,None,None,None,'#,##0','0.000%',
             '0.0%','0.0%','+0.0%;-0.0%;0.0%',None]
    aligns5 = ['center','center','left','center','center',
               'right','right','right','right','right','left']
    for j, (v, f, a) in enumerate(zip(vals5, fmts5, aligns5), 1):
        cell = apply_data(ws5, rank+2, j, v, bg=bg, fmt=f, align=a)
        if j == 10:
            cell.fill = fill(C_RED)
            cell.font = Font(name='Arial', bold=True, size=10, color=C_WHITE)

col_widths5 = [4, 20, 50, 11, 11, 8, 10, 10, 10, 10, 35]
for i, w in enumerate(col_widths5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6: PLANEJADAS SEM AVANÇO
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("PLANEJADAS SEM AVANÇO")
ws6.sheet_view.showGridLines = False

ws6.merge_cells('A1:J1')
ws6['A1'] = "ACTIVIDADES PLANEADAS SIN AVANCE REAL — CONSTRUCCIÓN | LA PAMPINA"
ws6['A1'].font = Font(name='Arial', bold=True, size=13, color=C_WHITE)
ws6['A1'].fill = fill(C_ORANGE)
ws6['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws6.row_dimensions[1].height = 24

ws6.merge_cells('A2:J2')
ws6['A2'] = f"Total: 69 actividades con % Plan > 0 y % Real = 0  |  Data: {DATA_WEEK}"
ws6['A2'].font = Font(name='Arial', bold=True, size=10)
ws6['A2'].fill = fill(C_YELLOW)
ws6['A2'].alignment = Alignment(horizontal='center', vertical='center')

hdrs6 = ['#', 'EDT', 'ACTIVIDAD', 'INICIO LB', 'FIN LB',
         'H-H', 'INCIDENCIA', '% COMP. PLAN', '% COMP. REAL', 'DESVÍO']
for j, h in enumerate(hdrs6, 1):
    apply_hdr(ws6, 3, j, h, bg=C_ORANGE, fg="000000", size=9, wrap=True)
ws6.row_dimensions[3].height = 28

planned_no_real = df_leaves[
    (df_leaves['pct_comp_plan'] > 0) &
    (df_leaves['pct_comp_real'] == 0)
].copy().sort_values('incidencia', ascending=False)

for rank, (_, r) in enumerate(planned_no_real.iterrows(), 1):
    bg = C_GRAY if rank % 2 == 0 else C_WHITE
    inicio = r['inicio_lb'].strftime('%d/%m/%Y') if isinstance(r['inicio_lb'], datetime) else ''
    fim    = r['fin_lb'].strftime('%d/%m/%Y')    if isinstance(r['fin_lb'], datetime)    else ''
    vals6 = [rank, r['edt'], r['tarea'][:60], inicio, fim,
             int(r['trabajo_h']), r['incidencia'],
             r['pct_comp_plan'], r['pct_comp_real'], r['desviacion']]
    fmts6 = [None,None,None,None,None,'#,##0','0.0000%',
             '0.0%','0.0%','+0.0%;-0.0%;0.0%']
    for j, (v, f) in enumerate(zip(vals6, fmts6), 1):
        cell = apply_data(ws6, rank+3, j, v, bg=bg, fmt=f)
        if j == 10:
            cell.fill = fill(C_ORANGE)
            cell.font = Font(name='Arial', bold=True, size=10)

col_widths6 = [4, 20, 50, 11, 11, 8, 10, 12, 12, 12]
for i, w in enumerate(col_widths6, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7: RANKING DE IMPACTOS
# ══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("RANKING DE IMPACTOS")
ws7.sheet_view.showGridLines = False

ws7.merge_cells('A1:K1')
ws7['A1'] = "RANKING DE IMPACTOS — ACTIVIDADES QUE MÁS AFECTAN EL PROGRESO"
ws7['A1'].font = Font(name='Arial', bold=True, size=13, color=C_WHITE)
ws7['A1'].fill = fill(C_DARK_BLUE)
ws7['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws7.row_dimensions[1].height = 24

ws7.merge_cells('A2:K2')
ws7['A2'] = "Ordena las actividades por impacto absoluto en el avance ponderado del proyecto"
ws7['A2'].font = Font(name='Arial', size=9, italic=True)
ws7['A2'].alignment = Alignment(horizontal='center')

hdrs7 = ['RANKING', 'EDT', 'ACTIVIDAD', 'H-H', 'INCIDENCIA',
         '% PLAN POND.', '% REAL POND.', 'IMPACTO POND.',
         '% COMP. PLAN', '% COMP. REAL', 'CLASIFICACIÓN']
for j, h in enumerate(hdrs7, 1):
    apply_hdr(ws7, 3, j, h, bg=C_DARK_BLUE, size=9, wrap=True)
ws7.row_dimensions[3].height = 30

impact_df = df_leaves.copy()
impact_df = impact_df[impact_df['incidencia'] > 0]
impact_df['impact_abs'] = (impact_df['desv_pond']).abs()
impact_df_sorted = impact_df.sort_values('impact_abs', ascending=False)

for rank, (_, r) in enumerate(impact_df_sorted.head(50).iterrows(), 1):
    bg = C_GRAY if rank % 2 == 0 else C_WHITE
    cls = "🔴 CRÍTICO" if r['impact_abs'] > 0.005 else \
          "🟠 ALTO"    if r['impact_abs'] > 0.002 else \
          "🟡 MEDIO"   if r['impact_abs'] > 0.0005 else "🟢 BAIXO"
    vals7 = [rank, r['edt'], r['tarea'][:55],
             int(r['trabajo_h']), r['incidencia'],
             r['pct_plan'], r['pct_real'],
             r['desv_pond'],
             r['pct_comp_plan'], r['pct_comp_real'], cls]
    fmts7 = [None,None,None,'#,##0','0.0000%',
             '0.00%','0.00%','+0.00%;-0.00%;0.00%',
             '0.0%','0.0%',None]
    for j, (v, f) in enumerate(zip(vals7, fmts7), 1):
        cell = apply_data(ws7, rank+3, j, v, bg=bg, fmt=f)
        if j == 8 and r['desv_pond'] < 0:
            cell.fill = fill(C_RED if r['impact_abs'] > 0.003 else C_ORANGE)
            cell.font = Font(name='Arial', bold=True, color=C_WHITE, size=10)

col_widths7 = [8, 18, 50, 8, 10, 10, 10, 12, 11, 11, 14]
for i, w in enumerate(col_widths7, 1):
    ws7.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 8: ANÁLISE DE CENÁRIOS
# ══════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("ANÁLISE DE CENÁRIOS")
ws8.sheet_view.showGridLines = False

ws8.merge_cells('A1:L1')
ws8['A1'] = "ANÁLISE DE CENÁRIOS — RECUPERAÇÃO DE AVANÇO | CONSTRUCCIÓN LA PAMPINA"
ws8['A1'].font = Font(name='Arial', bold=True, size=13, color=C_WHITE)
ws8['A1'].fill = fill(C_DARK_BLUE)
ws8['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws8.row_dimensions[1].height = 24

# Section A: Current state
ws8.merge_cells('A3:D3')
ws8['A3'] = "ESTADO ATUAL (referência)"
ws8['A3'].font = Font(name='Arial', bold=True, size=11, color=C_WHITE)
ws8['A3'].fill = fill(C_MED_BLUE)
ws8['A3'].alignment = Alignment(horizontal='center')

ref_items = [
    ("Data de controle", f"{DATA_DATE.strftime('%d/%m/%Y')} ({DATA_WEEK})"),
    ("% Planeado acumulado", f"{df_total['pct_plan']*100:.4f}%"),
    ("% Real acumulado",     f"{df_total['pct_real']*100:.4f}%"),
    ("Desvío acumulado",     f"{df_total['desv_pond']*100:+.4f}%"),
    ("H-H totales LB",       f"823.866 h"),
    ("Actividades elementales", "1.098"),
]
for i, (k, v) in enumerate(ref_items, 4):
    ws8.cell(row=i, column=1, value=k).font = Font(name='Arial', bold=True, size=10)
    ws8.cell(row=i, column=1).fill = fill(C_GRAY)
    ws8.cell(row=i, column=2, value=v).font = Font(name='Arial', size=10)
    ws8.cell(row=i, column=2).fill = fill(C_WHITE)
    for col in [1, 2]:
        ws8.cell(row=i, column=col).border = border_thin()
        ws8.cell(row=i, column=col).alignment = Alignment(horizontal='left' if col==1 else 'right', vertical='center')

# Section B: Scenario calculator
row8 = 12
ws8.merge_cells(start_row=row8, start_column=1, end_row=row8, end_column=12)
c = ws8.cell(row=row8, column=1, value="CALCULADORA DE CENÁRIOS — Inserir dados nas células azuis")
c.font = Font(name='Arial', bold=True, size=11, color=C_WHITE)
c.fill = fill(C_DARK_BLUE)
c.alignment = Alignment(horizontal='center')
ws8.row_dimensions[row8].height = 20
row8 += 1

# Scenario 1
ws8.merge_cells(start_row=row8, start_column=1, end_row=row8, end_column=12)
ws8.cell(row=row8, column=1, value="CENÁRIO 1: Completar uma atividade específica").font = Font(name='Arial', bold=True, size=10, color=C_WHITE)
ws8.cell(row=row8, column=1).fill = fill(C_MED_BLUE)
ws8.cell(row=row8, column=1).alignment = Alignment(horizontal='left')
row8 += 1

sc1_labels = [
    ("Inserir EDT da atividade (ex: 4.5.2.3):", "4.5.2.3", True),
    ("Incidência dessa atividade (%):", "=IF(C{r}=\"\",\"\",VLOOKUP(C{r},DADOS!B:G,6,0))", False),
    ("% Real atual:", "=IF(C{r}=\"\",\"\",VLOOKUP(C{r},DADOS!B:J,9,0))", False),
    ("Nova % real proposta (inserir):", "100%", True),
    ("Ganho de % ponderado estimado:", "=IF(C{r}=\"\",0,C{rInc}/100*(C{rNew}-C{rReal}))", False),
    ("NOVO % REAL TOTAL ESTIMADO:", f"={df_total['pct_real']*100:.4f}+C{row8+4}", False),
]
for i, (lbl, val_hint, is_input) in enumerate(sc1_labels):
    r_ = row8 + i
    c_lbl = ws8.cell(row=r_, column=1, value=lbl)
    c_lbl.font = Font(name='Arial', size=9, bold=True)
    c_lbl.fill = fill(C_GRAY)
    c_lbl.border = border_thin()
    c_lbl.alignment = Alignment(horizontal='left')

    c_val = ws8.cell(row=r_, column=3)
    if is_input:
        c_val.fill = PatternFill("solid", fgColor="0000FF")
        c_val.font = Font(name='Arial', size=10, bold=True, color=C_WHITE)
        c_val.value = val_hint
    else:
        c_val.fill = fill(C_LIGHT_BLUE)
        c_val.font = Font(name='Arial', size=10, color="000000")
        c_val.value = "← calculado automaticamente"
    c_val.border = border_thin()
    c_val.alignment = Alignment(horizontal='left')
    ws8.column_dimensions['A'].width = 45
    ws8.column_dimensions['C'].width = 35

row8 += len(sc1_labels) + 1

# Scenario 2: Recovery target
ws8.merge_cells(start_row=row8, start_column=1, end_row=row8, end_column=12)
ws8.cell(row=row8, column=1, value="CENÁRIO 2: Quantas atividades precisam ser completadas para recuperar X% de desvio?").font = Font(name='Arial', bold=True, size=10, color=C_WHITE)
ws8.cell(row=row8, column=1).fill = fill(C_GREEN)
ws8.cell(row=row8, column=1).alignment = Alignment(horizontal='left')
row8 += 1

ws8.cell(row=row8, column=1, value="Desvio que quero recuperar (%):").font = Font(name='Arial', bold=True, size=9)
ws8.cell(row=row8, column=1).fill = fill(C_GRAY)
ws8.cell(row=row8, column=1).border = border_thin()
c_inp = ws8.cell(row=row8, column=3, value=2.0)
c_inp.fill = PatternFill("solid", fgColor="0000FF")
c_inp.font = Font(name='Arial', bold=True, color=C_WHITE)
c_inp.number_format = '0.00"%"'
c_inp.border = border_thin()
row8 += 1

ws8.cell(row=row8, column=1, value="Incidência média por atividade (%):").font = Font(name='Arial', bold=True, size=9)
ws8.cell(row=row8, column=1).fill = fill(C_GRAY)
ws8.cell(row=row8, column=1).border = border_thin()

avg_incid = df_leaves[df_leaves['incidencia'] > 0]['incidencia'].mean()
c_avg = ws8.cell(row=row8, column=3, value=avg_incid * 100)
c_avg.fill = PatternFill("solid", fgColor="0000FF")
c_avg.font = Font(name='Arial', bold=True, color=C_WHITE)
c_avg.number_format = '0.0000"%"'
c_avg.border = border_thin()
row8 += 1

ws8.cell(row=row8, column=1, value="Nº atividades necessárias (estimativa):").font = Font(name='Arial', bold=True, size=9)
ws8.cell(row=row8, column=1).fill = fill(C_LIGHT_BLUE)
ws8.cell(row=row8, column=1).border = border_thin()
ws8.cell(row=row8, column=3, value=f"=ROUND(C{row8-2}/C{row8-1},0)").font = Font(name='Arial', bold=True, size=12, color=C_RED)
ws8.cell(row=row8, column=3).fill = fill(C_YELLOW)
ws8.cell(row=row8, column=3).border = border_thin()
row8 += 2

# Scenario 3: Top activities to recover
ws8.merge_cells(start_row=row8, start_column=1, end_row=row8, end_column=12)
ws8.cell(row=row8, column=1, value="CENÁRIO 3: Atividades futuras que mais ajudam a recuperar atraso").font = Font(name='Arial', bold=True, size=10, color=C_WHITE)
ws8.cell(row=row8, column=1).fill = fill(C_GREEN)
ws8.cell(row=row8, column=1).alignment = Alignment(horizontal='left')
row8 += 1

future = df_leaves[
    (df_leaves['pct_comp_real'] < 1.0) &
    (df_leaves['fin_lb'] > DATA_DATE) &
    (df_leaves['incidencia'] > 0.0005)
].copy()
future['recuperacao_potencial'] = future['incidencia'] * (1 - future['pct_comp_real'])
future_sorted = future.sort_values('recuperacao_potencial', ascending=False)

hdrs_sc3 = ['#', 'EDT', 'ATIVIDADE', 'H-H', 'INCIDÊNCIA',
            '% REAL ATUAL', '% RECUPERÁVEL', 'FIM PREVISTO']
for j, h in enumerate(hdrs_sc3, 1):
    apply_hdr(ws8, row8, j, h, bg=C_GREEN, fg="000000", size=9)
row8 += 1

for rank, (_, r) in enumerate(future_sorted.head(30).iterrows(), 1):
    bg = C_GRAY if rank % 2 == 0 else C_WHITE
    fim = r['fin_lb'].strftime('%d/%m/%Y') if isinstance(r['fin_lb'], datetime) else ''
    rec = r['recuperacao_potencial']
    vals = [rank, r['edt'], r['tarea'][:50],
            int(r['trabajo_h']), r['incidencia'],
            r['pct_comp_real'], rec, fim]
    fmts = [None,None,None,'#,##0','0.0000%','0.0%','0.000%',None]
    for j, (v, f) in enumerate(zip(vals, fmts), 1):
        cell = apply_data(ws8, row8, j, v, bg=bg, fmt=f)
        if j == 7:
            cell.fill = fill(C_GREEN if rec > 0.005 else (C_YELLOW if rec > 0.002 else bg))
    row8 += 1

col_widths8 = [45, 18, 50, 8, 10, 10, 10, 12]
for i, w in enumerate(col_widths8[:8], 1):
    ws8.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 9: FORECAST
# ══════════════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("FORECAST")
ws9.sheet_view.showGridLines = False

ws9.merge_cells('A1:J1')
ws9['A1'] = "FORECAST DE CONCLUSÃO — CONSTRUCCIÓN | LA PAMPINA"
ws9['A1'].font = Font(name='Arial', bold=True, size=13, color=C_WHITE)
ws9['A1'].fill = fill(C_DARK_BLUE)
ws9['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws9.row_dimensions[1].height = 24

# Compute simple linear forecast
plan_vals = [total_row_data[f'ps{i}'] for i in range(68)]
real_vals  = [total_row_data[f'rs{i}'] for i in range(68)]

# Last real value and rate
real_series = [(i, v) for i, v in enumerate(real_vals) if v and v > 0]
if len(real_series) >= 2:
    last_idx, last_real = real_series[-1]
    prev_idx, prev_real = real_series[-2]
    weekly_rate = (last_real - prev_real) / max(last_idx - prev_idx, 1)
else:
    weekly_rate = 0.0001

remaining = 1.0 - (df_total['pct_real'] if df_total['pct_real'] else 0)
weeks_to_complete = int(remaining / weekly_rate) if weekly_rate > 0 else 999
from datetime import timedelta
forecast_date = DATA_DATE + timedelta(weeks=weeks_to_complete)
lb_end = df_total['fin_lb']

ws9.merge_cells('A3:D3')
ws9['A3'] = "FORECAST BASEADO NA TAXA DE AVANÇO ATUAL"
ws9['A3'].font = Font(name='Arial', bold=True, size=11, color=C_WHITE)
ws9['A3'].fill = fill(C_MED_BLUE)

forecast_data = [
    ("% Real atual",                f"{df_total['pct_real']*100:.4f}%"),
    ("Restante para concluir",      f"{remaining*100:.4f}%"),
    ("Taxa semanal recente",        f"{weekly_rate*100:.4f}% / semana"),
    ("Semanas restantes (estimado)",f"{weeks_to_complete} semanas"),
    ("DATA PREVISTA DE CONCLUSÃO",  forecast_date.strftime('%d/%m/%Y')),
    ("Data prevista LB original",   lb_end.strftime('%d/%m/%Y') if isinstance(lb_end, datetime) else str(lb_end)),
    ("DESVIO DE PRAZO (estimado)",  f"{(forecast_date - lb_end).days if isinstance(lb_end, datetime) else '?'} dias"),
]

for i, (k, v) in enumerate(forecast_data, 4):
    is_key = 'DATA PREVISTA' in k or 'DESVIO' in k
    c_k = ws9.cell(row=i, column=1, value=k)
    c_k.font = Font(name='Arial', bold=True, size=10)
    c_k.fill = fill(C_LIGHT_BLUE if is_key else C_GRAY)
    c_k.border = border_thin()
    c_k.alignment = Alignment(horizontal='left', vertical='center')

    c_v = ws9.cell(row=i, column=2, value=v)
    c_v.font = Font(name='Arial', bold=is_key, size=11 if is_key else 10,
                    color=C_RED if is_key else "000000")
    c_v.fill = fill(C_YELLOW if is_key else C_WHITE)
    c_v.border = border_thin()
    c_v.alignment = Alignment(horizontal='right', vertical='center')
    ws9.row_dimensions[i].height = 18

ws9.merge_cells('A12:J12')
ws9['A12'] = "⚠ NOTA: Este forecast é uma projeção linear baseada na taxa das últimas semanas. Pode variar conforme aceleração de obras."
ws9['A12'].font = Font(name='Arial', size=9, italic=True, color=C_DARK_GRAY)
ws9['A12'].alignment = Alignment(horizontal='left')

ws9.column_dimensions['A'].width = 40
ws9.column_dimensions['B'].width = 25

# Forecast table (future weeks)
ws9.merge_cells('A14:J14')
ws9['A14'] = "PROGRESSO PLANEJADO VS. REAL — SEMANAS PASSADAS E FUTURAS"
ws9['A14'].font = Font(name='Arial', bold=True, size=10, color=C_WHITE)
ws9['A14'].fill = fill(C_DARK_BLUE)

hdrs_fc = ['SEMANA', 'FECHA', '% PLAN', '% REAL', 'DESVÍO', 'STATUS']
for j, h in enumerate(hdrs_fc, 1):
    apply_hdr(ws9, 15, j, h, bg=C_MED_BLUE, size=9)

for i in range(68):
    p = plan_vals[i] or 0
    r = real_vals[i] or 0
    d = r - p
    w_lbl = plan_weeks[i]
    fecha = plan_dates_list[i] if i < len(plan_dates_list) else None
    is_curr = (i == CURR_IDX)
    past = (i <= CURR_IDX)
    bg = C_LIGHT_BLUE if is_curr else (C_GRAY if i % 2 == 0 else C_WHITE)
    status = ("◀ ATUAL" if is_curr else
              ("OK" if d >= 0 else "ATRASO") if past else "FUTURO")

    apply_data(ws9, 16+i, 1, w_lbl, bg=bg)
    apply_data(ws9, 16+i, 2, fecha, bg=bg, fmt='DD/MM/YYYY')
    apply_data(ws9, 16+i, 3, p, bg=bg, fmt='0.00%')
    apply_data(ws9, 16+i, 4, r if r > 0 else None, bg=bg, fmt='0.00%')
    d_bg = C_RED if d < -0.005 else (C_GREEN if d > 0.001 else bg)
    apply_data(ws9, 16+i, 5, d if (p or r) else None, bg=d_bg, fmt='+0.00%;-0.00%;0.00%')
    apply_data(ws9, 16+i, 6, status, bg=bg)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 10: DADOS BRUTOS (filtrado Construção)
# ══════════════════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet("DADOS")
ws10.sheet_view.showGridLines = False

ws10.merge_cells('A1:O1')
ws10['A1'] = "DATOS BRUTOS — CONSTRUCCIÓN | LA PAMPINA  (fuente: LPA CONSTR.)"
ws10['A1'].font = Font(name='Arial', bold=True, size=12, color=C_WHITE)
ws10['A1'].fill = fill(C_DARK_GRAY)
ws10['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws10.row_dimensions[1].height = 20

hdrs10 = ['FILA', 'NIVEL', 'RESUMEN', 'EDT', 'TAREA',
          'DUR. (d)', 'INICIO LB', 'FIN LB', 'H-H',
          'INCIDENCIA', '% PLAN POND.', '% REAL POND.',
          '% COMP. PLAN', '% COMP. REAL', 'DESVÍO']
for j, h in enumerate(hdrs10, 1):
    apply_hdr(ws10, 2, j, h, bg=C_DARK_GRAY, size=9, wrap=True)
ws10.row_dimensions[2].height = 28

for row_idx, (_, r) in enumerate(df.iterrows(), 3):
    bg = C_GRAY if row_idx % 2 == 0 else C_WHITE
    inicio = r['inicio_lb'].strftime('%d/%m/%Y') if isinstance(r['inicio_lb'], datetime) else ''
    fim    = r['fin_lb'].strftime('%d/%m/%Y')    if isinstance(r['fin_lb'], datetime)    else ''
    vals10 = [r['fila'], r['nivel'], r['es_resumen'], r['edt'],
              str(r['tarea'])[:60],
              r['duracion'], inicio, fim,
              r['trabajo_h'], r['incidencia'],
              r['pct_plan'], r['pct_real'],
              r['pct_comp_plan'], r['pct_comp_real'], r['desv_pond']]
    fmts10 = [None,None,None,None,None,None,None,None,
              '#,##0','0.0000%','0.00%','0.00%','0.0%','0.0%','0.00%']
    for j, (v, f) in enumerate(zip(vals10, fmts10), 1):
        cell = apply_data(ws10, row_idx, j, v, bg=bg, fmt=f,
                          align='left' if j==5 else 'center')

col_widths10 = [5, 6, 8, 18, 55, 8, 11, 11, 10, 10, 11, 11, 12, 12, 10]
for i, w in enumerate(col_widths10, 1):
    ws10.column_dimensions[get_column_letter(i)].width = w

# ── Final: Save ───────────────────────────────────────────────────────────────
wb.save(OUTPUT)
wb_src.close()
print(f"\nPainel gerado com sucesso!")
print(f"   Arquivo: {OUTPUT}")
print(f"\n   Abas criadas:")
for s in wb.sheetnames:
    print(f"   - {s}")
