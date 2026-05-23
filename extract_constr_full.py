import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime

FILE = r"D:\PowerChina\Anexo 01-LPA.xlsm"
wb = load_workbook(FILE, read_only=True, keep_vba=True, data_only=True)
ws = wb["LPA CONSTR."]

all_rows = list(ws.iter_rows(values_only=True))

# Row 10 (index 9) = header
# Row 11 (index 10) = totals (Construction level)
# Rows 12+ (index 11+) = activities

# Column mapping (0-indexed):
# 0: N Fila, 1: Nivel, 2: Resumen, 3: SubProj, 4: Capitulo, 5: Id
# 6: EDT, 7: Tarea, 8: Duración, 9: Comienzo LB, 10: Fin LB, 11: Trabajo(h)
# 13: Incidencia
# 15: % Plan ponderado, 16: % Real ponderado
# 17: % Completado Plan, 18: % Completado Real
# 19: Desviacion Ponderada, 20: Desviacion
# 22-89: PLAN semanais (S44-25 a S07-27 = 68 cols)
# 92-159: REAL semanais (S44-25 a S07-27 = 68 cols)

# Get dates from row 9 (index 8)
date_row = all_rows[8]
plan_dates = []
for c in range(22, 90):
    v = date_row[c] if c < len(date_row) else None
    plan_dates.append(v)

# Current week is S20-26 = col 50 (PLAN), col 120 (REAL)
CURRENT_PLAN_COL = 50
CURRENT_REAL_COL = 120
DATA_DATE = datetime(2026, 5, 17)

records = []
for row in all_rows[10:]:  # skip header rows (0-9) and start from row 11
    if row[0] is None and row[5] is None:
        continue
    r = {
        'fila': row[0],
        'nivel': row[1],
        'es_resumen': row[2],
        'subproj': row[3],
        'capitulo': row[4],
        'id': row[5],
        'edt': row[6],
        'tarea': row[7],
        'duracion': row[8],
        'inicio_lb': row[9],
        'fin_lb': row[10],
        'trabajo_h': row[11],
        'incidencia': row[13] if len(row) > 13 else None,
        'pct_plan_ponderado': row[15] if len(row) > 15 else None,
        'pct_real_ponderado': row[16] if len(row) > 16 else None,
        'pct_completado_plan': row[17] if len(row) > 17 else None,
        'pct_completado_real': row[18] if len(row) > 18 else None,
        'desviacion_pond': row[19] if len(row) > 19 else None,
        'desviacion': row[20] if len(row) > 20 else None,
    }
    # PLAN weekly series
    for i, c in enumerate(range(22, 90)):
        r[f'plan_s{i}'] = row[c] if c < len(row) else None
    # REAL weekly series
    for i, c in enumerate(range(92, 160)):
        r[f'real_s{i}'] = row[c] if c < len(row) else None
    records.append(r)

df = pd.DataFrame(records)

# Save a simplified view
cols_meta = ['fila','nivel','es_resumen','edt','tarea','duracion','inicio_lb','fin_lb',
             'trabajo_h','incidencia','pct_plan_ponderado','pct_real_ponderado',
             'pct_completado_plan','pct_completado_real','desviacion_pond','desviacion']
df_meta = df[cols_meta].copy()

print("=== RESUMO GERAL DA CONSTRUÇÃO ===")
total_row = df_meta[df_meta['edt'] == '4.5']
if not total_row.empty:
    tr = total_row.iloc[0]
    print(f"Projeto: LA PAMPINA - Disciplina: CONSTRUÇÃO")
    print(f"Início LB: {tr['inicio_lb']}")
    print(f"Fim LB:    {tr['fin_lb']}")
    print(f"Trabalho Total (h): {tr['trabajo_h']:,.0f}")
    print(f"% Planejado Ponderado: {tr['pct_plan_ponderado']*100:.2f}%")
    print(f"% Real Ponderado:      {tr['pct_real_ponderado']*100:.2f}%")
    print(f"Desvio Acumulado:      {tr['desviacion_pond']*100:.2f}%")
    print(f"Data de controle:      {DATA_DATE.strftime('%d/%m/%Y')} (S20-26)")

print("\n=== ATIVIDADES NÍVEL 3 (Grupos Principais) ===")
nivel3 = df_meta[df_meta['nivel'] == 3].copy()
for _, r in nivel3.iterrows():
    plan = r['pct_plan_ponderado'] or 0
    real = r['pct_real_ponderado'] or 0
    dev = r['desviacion_pond'] or 0
    print(f"EDT: {r['edt']} | {str(r['tarea']).strip()}")
    print(f"   Início: {r['inicio_lb']} | Fim: {r['fin_lb']}")
    print(f"   Plan: {plan*100:.2f}% | Real: {real*100:.2f}% | Desvio: {dev*100:.2f}%")
    print(f"   Incidência: {(r['incidencia'] or 0)*100:.3f}% | H-H: {(r['trabajo_h'] or 0):,.0f}h")
    print()

print("\n=== TODAS AS ATIVIDADES (NÃO-RESUMO) ===")
leaves = df_meta[df_meta['es_resumen'] == 'No'].copy()
print(f"Total de atividades elementares: {len(leaves)}")

print("\n=== TOP 20 DESVIOS MAIS NEGATIVOS ===")
desvios = leaves.dropna(subset=['desviacion_pond']).copy()
desvios = desvios[desvios['incidencia'].notna() & (desvios['incidencia'] > 0)]
desvios_sorted = desvios.sort_values('desviacion_pond', ascending=True)
for _, r in desvios_sorted.head(20).iterrows():
    plan = r['pct_completado_plan'] or 0
    real = r['pct_completado_real'] or 0
    dev = r['desviacion'] or 0
    print(f"EDT: {r['edt']} | {str(r['tarea']).strip()[:60]}")
    print(f"   Plan%: {plan*100:.1f}% | Real%: {real*100:.1f}% | Desvio: {dev*100:.1f}%")
    print(f"   Incid.: {(r['incidencia'] or 0)*100:.4f}% | H-H: {(r['trabajo_h'] or 0):,.0f}h")

print("\n=== ATIVIDADES PLANEJADAS SEM AVANÇO REAL ===")
planned_no_real = leaves[
    (leaves['pct_completado_plan'].notna()) &
    (leaves['pct_completado_plan'] > 0) &
    (leaves['pct_completado_real'].notna()) &
    (leaves['pct_completado_real'] == 0)
].copy()
print(f"Quantidade: {len(planned_no_real)}")
for _, r in planned_no_real.iterrows():
    plan = r['pct_completado_plan'] or 0
    print(f"  EDT: {r['edt']} | {str(r['tarea']).strip()[:60]} | Plan: {plan*100:.1f}%")

print("\n=== ATIVIDADES COM PLANO FUTURO (>0 nas próximas semanas) ===")
future_activities = leaves[
    (leaves['pct_completado_real'].notna()) &
    (leaves['pct_completado_real'] < 1) &
    (leaves['fin_lb'].notna()) &
    (leaves['fin_lb'] > DATA_DATE)
].copy()
print(f"Atividades futuras/em andamento: {len(future_activities)}")

print("\n=== ATIVIDADES CRÍTICAS (Incidência alta e desvio negativo) ===")
if 'incidencia' in desvios_sorted.columns:
    critical = desvios_sorted[
        (desvios_sorted['incidencia'] > 0.005) &
        (desvios_sorted['desviacion_pond'] < -0.002)
    ].copy()
    critical_sorted = critical.sort_values('incidencia', ascending=False)
    for _, r in critical_sorted.head(15).iterrows():
        print(f"  EDT: {r['edt']} | {str(r['tarea']).strip()[:55]}")
        print(f"     Incid: {(r['incidencia'] or 0)*100:.4f}% | Desv.Pond.: {(r['desviacion_pond'] or 0)*100:.4f}%")

print("\n=== DISTRIBUIÇÃO POR CAPÍTULO/ÁREA ===")
nivel_counts = df_meta[df_meta['es_resumen'] == 'Sí'].groupby('nivel').size()
print("Grupos resumo por nível:", nivel_counts.to_dict())

print("\n=== SEMANAS DE DADOS DISPONÍVEIS ===")
header_row = all_rows[9]  # row 10 (0-indexed = 9)
plan_week_headers = []
for c in range(22, 90):
    v = header_row[c] if c < len(header_row) else None
    if v: plan_week_headers.append((c, v))
print(f"Semanas PLAN: {plan_week_headers[0]} até {plan_week_headers[-1]}")

real_week_headers = []
for c in range(92, 160):
    v = header_row[c] if c < len(header_row) else None
    if v: real_week_headers.append((c, v))
print(f"Semanas REAL: {real_week_headers[0]} até {real_week_headers[-1]}")

# Save full extracted data
df_meta.to_csv(r"D:\PowerChina\constr_activities.csv", index=False, encoding='utf-8-sig')
print("\n\nDados salvos em: D:\\PowerChina\\constr_activities.csv")

wb.close()
