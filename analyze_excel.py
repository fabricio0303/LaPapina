import openpyxl
import pandas as pd
from openpyxl import load_workbook
from collections import Counter
import json
import warnings
warnings.filterwarnings('ignore')

FILE_PATH = r"D:\PowerChina\extracted\Anexo 02 Progreso Fisico.xlsm"

print("=" * 80)
print("STEP 1: LOADING WORKBOOK AND LISTING SHEETS")
print("=" * 80)

wb = load_workbook(FILE_PATH, read_only=True, keep_vba=True, data_only=True)
sheet_names = wb.sheetnames
print(f"Total sheets: {len(sheet_names)}")
for i, name in enumerate(sheet_names):
    print(f"  [{i+1}] {name}")

wb.close()

print("\n" + "=" * 80)
print("STEP 2: PER-SHEET DIMENSION AND HEADER ANALYSIS")
print("=" * 80)

results = {}

for sheet_name in sheet_names:
    print(f"\n--- Sheet: '{sheet_name}' ---")
    try:
        wb2 = load_workbook(FILE_PATH, read_only=True, keep_vba=True, data_only=True)
        ws = wb2[sheet_name]

        max_row = ws.max_row
        max_col = ws.max_column
        print(f"  Dimensions: {max_row} rows x {max_col} columns")

        # Read first 10 rows to find header row
        rows_data = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(15, max_row or 15), values_only=True)):
            rows_data.append(row)

        # Find header row - look for row with most non-None values or containing key words
        header_row_idx = None
        header_row = None
        key_words = ['task', 'name', 'nombre', 'actividad', 'descripcion', 'descripción',
                     'edt', 'wbs', 'inicio', 'fin', 'start', 'finish', 'duration',
                     'progreso', 'avance', 'peso', 'incidencia', 'subcontrat',
                     'area', 'área', 'disciplina', 'estado', 'status']

        best_score = 0
        for idx, row in enumerate(rows_data):
            non_none = sum(1 for v in row if v is not None)
            score = non_none
            row_str = ' '.join(str(v).lower() for v in row if v is not None)
            for kw in key_words:
                if kw in row_str:
                    score += 5
            if score > best_score:
                best_score = score
                header_row_idx = idx + 1
                header_row = row

        print(f"  Best header row candidate: Row {header_row_idx}")
        if header_row:
            headers = [(i+1, str(v) if v is not None else 'None') for i, v in enumerate(header_row) if v is not None]
            print(f"  Headers found ({len(headers)} non-None):")
            for col_num, h in headers:
                print(f"    Col {col_num}: {h}")

        results[sheet_name] = {
            'max_row': max_row,
            'max_col': max_col,
            'header_row_idx': header_row_idx,
            'headers': header_row
        }
        wb2.close()
    except Exception as e:
        print(f"  ERROR: {e}")
        results[sheet_name] = {'error': str(e)}

print("\n" + "=" * 80)
print("STEP 3: DETAILED ANALYSIS OF MAIN DATA SHEET")
print("=" * 80)

# Try to read each sheet with pandas and find the main one
main_sheet = None
main_df = None

for sheet_name in sheet_names:
    try:
        # Try different header rows
        for header_row in [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]:
            try:
                df = pd.read_excel(FILE_PATH, sheet_name=sheet_name, header=header_row, engine='openpyxl')
                col_str = ' '.join(str(c).lower() for c in df.columns)
                # Check if this looks like a data sheet
                score = 0
                for kw in ['task', 'name', 'nombre', 'actividad', 'edt', 'wbs', 'inicio', 'start',
                           'progreso', 'avance', 'peso', 'subcontrat']:
                    if kw in col_str:
                        score += 1
                if score >= 2 and len(df) > 5:
                    if main_df is None or len(df) > len(main_df):
                        main_sheet = sheet_name
                        main_df = df
                        main_header_row = header_row
                    break
            except:
                pass
    except Exception as e:
        print(f"  Skipping {sheet_name}: {e}")

if main_sheet is None:
    # Fallback: use the sheet with most rows
    max_rows = 0
    for sname, sdata in results.items():
        if 'max_row' in sdata and sdata['max_row'] and sdata['max_row'] > max_rows:
            max_rows = sdata['max_row']
            main_sheet = sname

print(f"\nIdentified main data sheet: '{main_sheet}'")

if main_df is not None:
    print(f"  Shape: {main_df.shape[0]} rows x {main_df.shape[1]} columns")
    print(f"  Header row used: {main_header_row}")
    print(f"\n  ALL COLUMN HEADERS:")
    for i, col in enumerate(main_df.columns):
        dtype = str(main_df[col].dtype)
        non_null = main_df[col].notna().sum()
        pct = 100 * non_null / len(main_df) if len(main_df) > 0 else 0
        print(f"    [{i+1:3d}] '{col}' | dtype: {dtype} | non-null: {non_null}/{len(main_df)} ({pct:.1f}%)")

print("\n" + "=" * 80)
print("STEP 4: SAMPLE DATA (first 30 rows)")
print("=" * 80)

if main_df is not None:
    sample = main_df.head(30)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_colwidth', 50)
    pd.set_option('display.width', 200)
    print(sample.to_string())

print("\n" + "=" * 80)
print("STEP 5: IDENTIFY KEY COLUMNS")
print("=" * 80)

if main_df is not None:
    cols_lower = {str(c).lower(): c for c in main_df.columns}

    edt_cols = [c for k, c in cols_lower.items() if any(x in k for x in ['edt', 'wbs', 'cod', 'id'])]
    task_cols = [c for k, c in cols_lower.items() if any(x in k for x in ['task', 'name', 'nombre', 'actividad', 'descripcion', 'descripción'])]
    date_cols = [c for k, c in cols_lower.items() if any(x in k for x in ['date', 'fecha', 'inicio', 'fin', 'start', 'finish', 'comienzo'])]
    progress_cols = [c for k, c in cols_lower.items() if any(x in k for x in ['progreso', 'avance', 'progress', '%', 'porcentaje', 'percent'])]
    weight_cols = [c for k, c in cols_lower.items() if any(x in k for x in ['peso', 'incidencia', 'weight', 'ponderacion', 'ponderación'])]
    subcon_cols = [c for k, c in cols_lower.items() if any(x in k for x in ['subcontrat', 'contratist', 'contractor', 'empresa'])]
    area_cols = [c for k, c in cols_lower.items() if any(x in k for x in ['area', 'área', 'disciplina', 'discipline', 'especialidad'])]
    status_cols = [c for k, c in cols_lower.items() if any(x in k for x in ['estado', 'status', 'estatus', 'situacion', 'situación'])]

    print(f"  EDT/WBS/Code columns: {edt_cols}")
    print(f"  Task/Name columns: {task_cols}")
    print(f"  Date columns: {date_cols}")
    print(f"  Progress % columns: {progress_cols}")
    print(f"  Weight/Incidence columns: {weight_cols}")
    print(f"  Subcontractor columns: {subcon_cols}")
    print(f"  Area/Discipline columns: {area_cols}")
    print(f"  Status columns: {status_cols}")

print("\n" + "=" * 80)
print("STEP 6: UNIQUE VALUES IN CATEGORICAL COLUMNS")
print("=" * 80)

if main_df is not None:
    cat_candidate_cols = area_cols + subcon_cols + status_cols
    # Also add low-cardinality string columns
    for col in main_df.columns:
        if main_df[col].dtype == object:
            nuniq = main_df[col].nunique()
            if 1 < nuniq <= 50 and col not in cat_candidate_cols:
                cat_candidate_cols.append(col)

    for col in cat_candidate_cols[:20]:  # limit to 20 columns
        try:
            vals = main_df[col].dropna().unique().tolist()
            vals_str = [str(v) for v in vals]
            print(f"\n  Column '{col}' ({len(vals)} unique values):")
            for v in sorted(vals_str)[:30]:
                count = (main_df[col].astype(str) == v).sum()
                print(f"    - '{v}': {count} rows")
        except Exception as e:
            print(f"  Column '{col}': ERROR - {e}")

print("\n" + "=" * 80)
print("STEP 7: DATE RANGES")
print("=" * 80)

if main_df is not None:
    for col in main_df.columns:
        try:
            if 'date' in str(col).lower() or 'fecha' in str(col).lower() or \
               'inicio' in str(col).lower() or 'fin' in str(col).lower() or \
               'start' in str(col).lower() or 'finish' in str(col).lower() or \
               'comienzo' in str(col).lower():
                s = pd.to_datetime(main_df[col], errors='coerce')
                valid = s.dropna()
                if len(valid) > 0:
                    print(f"  Column '{col}':")
                    print(f"    Min: {valid.min()}")
                    print(f"    Max: {valid.max()}")
                    print(f"    Non-null: {len(valid)}/{len(main_df)}")
        except:
            pass

    # Also check all datetime-typed columns
    for col in main_df.select_dtypes(include=['datetime64']).columns:
        valid = main_df[col].dropna()
        if len(valid) > 0:
            print(f"  Column '{col}' (datetime):")
            print(f"    Min: {valid.min()}")
            print(f"    Max: {valid.max()}")
            print(f"    Non-null: {len(valid)}/{len(main_df)}")

print("\n" + "=" * 80)
print("STEP 8: MISSING DATA PATTERNS")
print("=" * 80)

if main_df is not None:
    total = len(main_df)
    print(f"  Total rows: {total}")
    missing = main_df.isnull().sum()
    missing_pct = (missing / total * 100).round(1)
    missing_df = pd.DataFrame({'missing_count': missing, 'missing_pct': missing_pct})
    missing_df = missing_df[missing_df['missing_count'] > 0].sort_values('missing_pct', ascending=False)
    if len(missing_df) > 0:
        print(f"\n  Columns with missing data ({len(missing_df)} columns):")
        for col, row in missing_df.iterrows():
            print(f"    '{col}': {int(row['missing_count'])} missing ({row['missing_pct']}%)")
    else:
        print("  No missing data found!")

    # Completely empty rows
    empty_rows = main_df.isnull().all(axis=1).sum()
    print(f"\n  Completely empty rows: {empty_rows}")

print("\n" + "=" * 80)
print("STEP 9: NUMERIC COLUMN STATISTICS")
print("=" * 80)

if main_df is not None:
    num_cols = main_df.select_dtypes(include=['number']).columns
    print(f"  Numeric columns ({len(num_cols)}):")
    for col in num_cols:
        s = main_df[col].dropna()
        if len(s) > 0:
            print(f"  '{col}': min={s.min():.4f}, max={s.max():.4f}, mean={s.mean():.4f}, non-null={len(s)}")

print("\n" + "=" * 80)
print("STEP 10: FULL COLUMN NAMES FROM RAW OPENPYXL (ALL SHEETS)")
print("=" * 80)

wb3 = load_workbook(FILE_PATH, read_only=True, keep_vba=True, data_only=True)
for sheet_name in sheet_names:
    print(f"\n--- Sheet: '{sheet_name}' ---")
    ws = wb3[sheet_name]
    # Read first 12 rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
        non_none = [v for v in row if v is not None]
        if non_none:
            print(f"  Row {row_idx}: {row}")
wb3.close()

print("\n" + "=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)
