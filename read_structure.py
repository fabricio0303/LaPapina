import pandas as pd
import openpyxl
from openpyxl import load_workbook

FILE = r"D:\PowerChina\Anexo 01-LPA.xlsm"

# --- 1. List all sheets ---
wb = load_workbook(FILE, read_only=True, keep_vba=True, data_only=True)
print("=== ABAS ENCONTRADAS ===")
for name in wb.sheetnames:
    print(f"  - {name}")

print()

# --- 2. For each sheet, show first rows ---
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n{'='*60}")
    print(f"ABA: {name}")

    rows_data = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True)):
        rows_data.append(row)
        if i >= 7:
            break

    max_col = max((len(r) for r in rows_data), default=0)
    print(f"Colunas detectadas: {max_col}")
    print("Primeiras 8 linhas:")
    for i, row in enumerate(rows_data):
        non_none = [(j, v) for j, v in enumerate(row) if v is not None]
        print(f"  Linha {i+1}: {non_none[:20]}")

wb.close()
